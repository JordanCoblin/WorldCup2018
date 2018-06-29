import openpyxl
from os import listdir, path

excel_sheets = {
	"Jordan": "World Cup 2018 Sheet Jordan.xlsx",
	"Jeff" : "2018 WC Pool Jeff.xlsx",
	"Millie" : "World Cup 2018 Sheet - Emile.xlsx",
	"Nish" : "World Cup 2018 Sheet Nish.xlsx",
	"Shonah" : "World Cup 2018 Sheet Shonah.xlsx",
	"Patty" : "World Cup 2018 Sheet_Pfingler.xlsx",
	"Vineet" : "VINNY_WORLD_CUP.xlsx",
	"Darnel" : "MD World Cup 2018 Sheet.xlsx",
}

RESULTS_WB =  "Results.xlsx"
STATS_WB = "bro_stats.xlsx"

# Results are of the form []
def read_part4(workbook):
	wb = openpyxl.load_workbook(path.join("picks", workbook))
	ws = wb.active

	results = []
	for row in ws.iter_rows(min_row=69, max_row=116, min_col=3, max_col=5):
		for i, cell in enumerate(row):
			if str(cell.value).lower() == "x":
				results.append(i)
	return results

def read_part2(workbook):
	wb = openpyxl.load_workbook(path.join("picks", workbook))
	ws = wb.active

	results = []
	group_result_1 = []
	group_result_2 = []
	for i, row in enumerate(ws.iter_rows(min_row=18, max_row=37, min_col=3, max_col=6)):

		# Next group result
		if (i+1)%5 == 0:
			results.append(group_result_1)
			results.append(group_result_2)
			group_result_1 = []
			group_result_2 = []
			continue

		if str(row[0].value).lower() == "x":
			group_result_1.append((i)%5)

		if str(row[3].value).lower() == "x":
			group_result_2.append((i)%5)

	return results

def read_part3(workbook):
	wb = openpyxl.load_workbook(path.join("picks", workbook))
	ws = wb.active

	results = []
	group_result_1 = -1
	group_result_2 = -1
	for i, row in enumerate(ws.iter_rows(min_row=43, max_row=62, min_col=3, max_col=6)):

		# Next group result
		if (i+1)%5 == 0:
			results.append(group_result_1)
			results.append(group_result_2)
			group_result_1 = -1
			group_result_2 = -1
			continue

		if str(row[0].value).lower() == "x":
			group_result_1 = i%5

		if str(row[3].value).lower() == "x":
			group_result_2 = i%5

	return results

def calculate_part2_score(picks, results):
	score = 0
	for i, group_result in enumerate(results):
		for team in group_result:
			if team in picks[i]:
				score += 2
	return score

def calculate_part3_score(picks, results):
	score = 0
	for i, group_result in enumerate(results):
		if picks[i] == group_result:
			score += 3
	return score

def calculate_part4_score(picks, results):
	score = 0
	for i, result in enumerate(results):
		if picks[i] == result:
			score += 1
	return score

def write_stats(scores):
	wb = openpyxl.load_workbook(path.join("picks", STATS_WB))
	wb.remove(wb['Sheet1'])
	wb.create_sheet('Sheet1')

	ws = wb.active

	header_row = ["Bro", "Part2 Score", "Part3 Score", "Part4 Score", "Total"]
	ws.append(header_row)

	print(scores)
	for bro_name, part_scores in scores.items():
		row = [bro_name]
		for score in part_scores:
			row.append(int(score))

		ws.append(row)

	wb.save(path.join("picks", STATS_WB))

# Read from actual results
results_part2 = read_part2(RESULTS_WB)
results_part3 = read_part3(RESULTS_WB)
results_part4 = read_part4(RESULTS_WB)

bro_scores = {}
for bro, workbook in excel_sheets.items():
	bro_part2 = read_part2(workbook)
	bro_part3 = read_part3(workbook)
	bro_part4 = read_part4(workbook)

	bro_scores[bro] = []
	bro_part2_score = calculate_part2_score(bro_part2, results_part2)
	bro_part3_score = calculate_part3_score(bro_part3, results_part3)
	bro_part4_score = calculate_part4_score(bro_part4, results_part4)
	
	bro_scores[bro].append(bro_part2_score)
	bro_scores[bro].append(bro_part3_score)
	bro_scores[bro].append(bro_part4_score)
	bro_scores[bro].append(bro_part2_score + bro_part3_score + bro_part4_score)

print (bro_scores)
write_stats(bro_scores)